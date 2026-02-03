import os
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT_DIR = "."
OUTPUT_XLSX = "m_by_n_st_grid.xlsx"

FOLDER_RE = re.compile(r"paths_m(\d+)_n(\d+)")
FILE_RE = re.compile(
    r"paths_m\d+_n\d+_S\((\d+),(\d+)\)_T\((\d+),(\d+)\)\.txt"
)

RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
BOLD = Font(bold=True)


def parse_file(path):
    m = n = count = None
    S = T = None

    with open(path, "r") as f:
        for line in f:
            line = line.strip()
            if line.startswith("m="):
                m = int(line[2:])
            elif line.startswith("n="):
                n = int(line[2:])
            elif line.startswith("S="):
                S = line[2:]
            elif line.startswith("T="):
                T = line[2:]
            elif line.startswith("count="):
                count = int(line[6:])

    return m, n, S, T, count


def main():
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    ns_by_m = defaultdict(set)

    # read all files
    for folder in os.listdir(ROOT_DIR):
        folder_path = os.path.join(ROOT_DIR, folder)
        if not os.path.isdir(folder_path):
            continue
        if not FOLDER_RE.match(folder):
            continue

        for fname in os.listdir(folder_path):
            match = FILE_RE.match(fname)
            if not match:
                continue

            full_path = os.path.join(folder_path, fname)
            m, n, S, T, count = parse_file(full_path)

            Sx = int(S[1:-1].split(",")[0])
            Tx = int(T[1:-1].split(",")[0])

            data[m][n][Sx][Tx] = count
            ns_by_m[m].add(n)

    wb = Workbook()
    ws = wb.active
    ws.title = "Hamiltonian Paths"

    row_idx = 1

    for m in sorted(data.keys()):
        ns = sorted(ns_by_m[m])

        # Header row
        col_idx = 1
        for n in ns:
            headers = ["m", "n", "S", "T", "count"]
            for h in headers:
                cell = ws.cell(row=row_idx, column=col_idx, value=h)
                cell.font = BOLD
                col_idx += 1
            col_idx += 1  # empty column between n-blocks

        row_idx += 1

        # Data rows
        for Sx in range(m):
            for Tx in range(m):
                col_idx = 1
                for n in ns:
                    S = f"({Sx},0)"
                    T = f"({Tx},{n-1})"
                    count = data[m][n].get(Sx, {}).get(Tx, None)

                    values = [m, n, S, T, count if count is not None else ""]

                    fill = None
                    if count == 0:
                        fill = RED_FILL

                    for v in values:
                        cell = ws.cell(row=row_idx, column=col_idx, value=v)
                        if fill:
                            cell.fill = fill
                        col_idx += 1

                    col_idx += 1  # empty column

                row_idx += 1

        # blank line between m-blocks
        row_idx += 2

    wb.save(OUTPUT_XLSX)
    print(f"Excel written to {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
