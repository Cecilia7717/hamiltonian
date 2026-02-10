import os
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT_DIR = "."
OUTPUT_XLSX = "m_by_n_st_grid_special.xlsx"

FOLDER_RE = re.compile(r"paths_m(\d+)_n(\d+)")
FILE_RE = re.compile(
    r"paths_m\d+_n\d+_S\((\d+),(\d+)\)_T\((\d+),(\d+)\)\.txt"
)

RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
PURPLE_FILL = PatternFill(start_color="E0CCFF", end_color="E0CCFF", fill_type="solid")

BOLD = Font(bold=True)


def parse_file(path):
    m = n = count = None
    S = T = None

    with open(path, "r") as f:
        for line in f:
            line = line.strip()
            if line.startswith("m="):
                m = int(line[2:])
            elif line.startswith("n="):
                n = int(line[2:])
            elif line.startswith("S="):
                S = line[2:]
            elif line.startswith("T="):
                T = line[2:]
            elif line.startswith("count="):
                count = int(line[6:])

    return m, n, S, T, count


def theorem_allows_path(m, n, x1, x2):
    if m % 2 == 0:
        if m == 2:
            if n % 2 == 1:
                return (x1, x2) in [(0, 1), (1, 0)]
            else:
                return (x1, x2) in [(0, 0), (1, 1)]
        else:
            if n % 2 == 0:
                if n == 2:
                    return (
                        (x1 % 2 == x2 % 2)
                        and (
                            (x1 != x2)
                            or (x1 == x2 == 0)
                            or (x1 == x2 == m - 1)
                        )
                    )
                else:
                    return x1 % 2 == x2 % 2
            else:
                # ---------- n odd ----------
                if n == 3:
                    # parity condition
                    if x1 % 2 == x2 % 2:
                        return False

                    A = {0, 2, m - 3, m - 1}
                    if x1 in A:
                        return True

                    even_index = list(range(2, m, 2))      # {2,4,...,m-2}
                    odd_index  = list(range(1, m - 4, 2))  # {1,3,...,m-5}

                    if x1 % 2 == 1:
                        # x1 odd
                        k = (x1 + 1) // 2
                        forbidden = even_index[k:]
                        return x2 not in forbidden
                    else:
                        # x1 even
                        k = (m - 2 - x1) // 2
                        forbidden = odd_index[:len(odd_index) - k]
                        return x2 not in forbidden
                else:
                    return x1 % 2 != x2 % 2
    else:
        if m == 3:
            return (x1, x2) in [
                (0, 0),
                (m - 1, 0),
                (0, m - 1),
                (m - 1, m - 1),
            ]
        else:
            if n % 2 == 0:
                if n == 2:
                    return (
                        (x1 % 2 == x2 % 2)
                        and (
                            (x1 != x2)
                            or (x1 == x2 == 0)
                            or (x1 == x2 == m - 1)
                        )
                    )
                else:
                    return x1 % 2 == x2 % 2
            else:
                return (
                    x1 % 2 == 0
                    and x2 % 2 == 0
                )



def main():
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    ns_by_m = defaultdict(set)

    for folder in os.listdir(ROOT_DIR):
        folder_path = os.path.join(ROOT_DIR, folder)
        if not os.path.isdir(folder_path):
            continue
        if not FOLDER_RE.match(folder):
            continue

        for fname in os.listdir(folder_path):
            match = FILE_RE.match(fname)
            if not match:
                continue

            full_path = os.path.join(folder_path, fname)
            m, n, S, T, count = parse_file(full_path)

            Sx = int(S[1:-1].split(",")[0])
            Tx = int(T[1:-1].split(",")[0])

            data[m][n][Sx][Tx] = count
            ns_by_m[m].add(n)

    wb = Workbook()
    ws = wb.active
    ws.title = "Hamiltonian Paths"

    row_idx = 1

    for m in sorted(data.keys()):
        ns = sorted(ns_by_m[m])

        col_idx = 1
        for n in ns:
            for h in ["m", "n", "S", "T", "count"]:
                cell = ws.cell(row=row_idx, column=col_idx, value=h)
                cell.font = BOLD
                col_idx += 1
            col_idx += 1

        row_idx += 1

        for Sx in range(m):
            for Tx in range(m):
                col_idx = 1
                for n in ns:
                    count = data[m][n].get(Sx, {}).get(Tx, None)
                    allowed = theorem_allows_path(m, n, Sx, Tx)

                    fill = None
                    if count == 0 and not allowed:
                        fill = GREEN_FILL
                    elif count == 0 and allowed:
                        fill = RED_FILL
                    elif count and count > 0 and not allowed:
                        fill = PURPLE_FILL

                    values = [m, n, f"({Sx},0)", f"({Tx},{n-1})", count]

                    for v in values:
                        cell = ws.cell(row=row_idx, column=col_idx, value=v)
                        if fill:
                            cell.fill = fill
                        col_idx += 1

                    col_idx += 1

                row_idx += 1

        row_idx += 2

    wb.save(OUTPUT_XLSX)
    print(f"Excel written to {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
